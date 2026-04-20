from pathlib import Path

from openpyxl import load_workbook


def get_workbook_info(path: Path) -> list[str]:
    """Return list of sheet names."""
    wb = load_workbook(path, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def write_extraction_to_excel(
    wb_path: Path,
    sheet_name: str,
    mapping: dict[str, tuple[str, int]],  # key -> (column_letter, start_row)
    data: list[dict[str, str]],
    output_path: Path,
) -> None:
    """Write extraction results into an Excel file.

    mapping: {"Họ tên": ("B", 5), "Ngày sinh": ("C", 5)}
    data: [{"Họ tên": "Nguyễn Văn A", ...}, ...]
    """
    wb = load_workbook(wb_path)
    ws = wb[sheet_name]

    for i, row_data in enumerate(data):
        for key, (col, start_row) in mapping.items():
            value = row_data.get(key, "")
            cell = f"{col}{start_row + i}"
            ws[cell] = value

    wb.save(output_path)
    wb.close()
